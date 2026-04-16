#!/usr/bin/env python3
"""
xlsx-manager.py — SEO Contextualizador Squad
Lê e escreve na planilha XLSX + detecta/cria cliente automaticamente.

Uso:
  python xlsx-manager.py --action=read              --file=input.xlsx
  python xlsx-manager.py --action=detect-client     --file=input.xlsx
  python xlsx-manager.py --action=auto-onboard      --file=input.xlsx --context=contexto.txt
  python xlsx-manager.py --action=save-client       --slug=nome --nome="Nome" --context=contexto.txt [--aliases=a,b,c] [--setor="Setor"]
  python xlsx-manager.py --action=list-clients
  python xlsx-manager.py --action=write             --input=input.xlsx --output=output.xlsx --data=results.json

Auto-onboarding (primeiro acesso):
  Recebe o xlsx + o texto do contexto da empresa.
  Extrai o nome da empresa do contexto, verifica se já existe em clients/,
  e retorna: use_existing | create_new | confirm_match (nomes similares).
  O Maestro decide o que fazer com base no resultado.
"""

import sys
import json
import re
import argparse
from pathlib import Path

# Raiz do squad — dois níveis acima deste script (scripts/ → squad-root/)
SQUAD_ROOT = Path(__file__).resolve().parent.parent


# ---------------------------------------------------------------------------
# Mapeamento de headers da planilha
# ---------------------------------------------------------------------------

HEADER_MAP = {
    "idioma": "idioma",
    "palavra-chave foco": "keyword",
    "titulo": "titulo",
    "título": "titulo",
    "slug (url amigavel)": "slug",
    "slug (url amigável)": "slug",
    "post type": "post_type",
    "categoria": "categoria",
    "tags": "tags",
    "banco de imagem": "banco_imagem",
    "imagens nos subtitulos": "imagens_subtitulos",
    "imagens nos subtítulos": "imagens_subtitulos",
    "quantidade de subtitulos": "qtd_subtitulos",
    "quantidade de subtítulos": "qtd_subtitulos",
    "status": "status",
    "resumo": "resumo",
    "conclusao": "conclusao",
    "conclusão": "conclusao",
    "faq": "faq",
    "link interno": "link_interno",
    "video": "video",
    "vídeo": "video",
    "autor id (wordpress)": "autor_id",
    "termo do link": "termo_link",
    "url do link": "url_link",
    "prompt adicional": "prompt_adicional",
    "termos lsi": "termos_lsi",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("ERRO: openpyxl não instalado. Execute: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def normalize(text):
    if text is None:
        return ""
    return str(text).strip().lower()


def slugify(text):
    """Converte texto em slug para comparação: 'Mand Digital' → 'mand-digital'."""
    text = normalize(text)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def read_headers(ws):
    header_row = ws[1]
    col_map = {}
    col_names = {}
    for cell in header_row:
        if cell.value is not None:
            norm = normalize(str(cell.value))
            field = HEADER_MAP.get(norm, norm)
            col_map[field] = cell.column
            col_names[cell.column] = str(cell.value)
    return col_map, col_names


# ---------------------------------------------------------------------------
# Detecção de cliente
# ---------------------------------------------------------------------------

def load_client_index():
    """Carrega clients/_index.yaml e retorna lista de clientes."""
    index_path = SQUAD_ROOT / "clients" / "_index.yaml"
    if not index_path.exists():
        return []

    try:
        import yaml
        with open(index_path, encoding="utf-8") as f:
            data = yaml.safe_load(f)
        return data.get("clientes", []) if data else []
    except ImportError:
        # Fallback sem PyYAML: parse manual simples
        return _parse_index_without_yaml(index_path)
    except Exception:
        return []


def _parse_index_without_yaml(index_path):
    """Parser mínimo para _index.yaml sem PyYAML instalado."""
    clientes = []
    current = None
    in_aliases = False

    with open(index_path, encoding="utf-8") as f:
        for line in f:
            stripped = line.strip()

            if stripped.startswith("- slug:"):
                if current:
                    clientes.append(current)
                current = {"slug": stripped.split(":", 1)[1].strip(), "aliases": [], "ativo": True}
                in_aliases = False

            elif current and stripped.startswith("nome:"):
                current["nome"] = stripped.split(":", 1)[1].strip().strip('"').strip("'")

            elif current and stripped.startswith("arquivo:"):
                current["arquivo"] = stripped.split(":", 1)[1].strip().strip('"').strip("'")

            elif current and stripped.startswith("aliases:"):
                in_aliases = True

            elif current and in_aliases and stripped.startswith("- "):
                alias = stripped[2:].strip().strip('"').strip("'")
                current["aliases"].append(alias)

            elif current and stripped.startswith("ativo:"):
                current["ativo"] = "true" in stripped.lower()

            elif stripped and not stripped.startswith("#") and not stripped.startswith("-"):
                if ":" in stripped:
                    in_aliases = False

    if current:
        clientes.append(current)

    return clientes


def detect_client_from_filename(filename, clientes):
    """
    Tenta detectar o cliente a partir do nome do arquivo xlsx.
    Retorna: (cliente_dict | None, confianca: "alta" | "baixa" | "nenhuma")
    """
    stem = slugify(Path(filename).stem)   # ex: "mand-digital-posts-abril-2026"
    tokens = set(re.split(r"[-_\s]+", stem))  # {'mand', 'digital', 'posts', 'abril', '2026'}

    matches = []
    for cliente in clientes:
        if not cliente.get("ativo", True):
            continue
        for alias in cliente.get("aliases", []):
            alias_slug = slugify(alias)
            alias_tokens = set(re.split(r"[-_\s]+", alias_slug))
            # Match se todos os tokens do alias estão no nome do arquivo
            if alias_tokens and alias_tokens.issubset(tokens):
                matches.append((cliente, alias))
                break

    if len(matches) == 1:
        return matches[0][0], "alta"
    elif len(matches) > 1:
        # Múltiplos matches — retorna o mais específico (mais tokens no alias)
        matches.sort(key=lambda m: len(m[1]), reverse=True)
        return matches[0][0], "baixa"
    return None, "nenhuma"


def load_client_context(cliente):
    """Lê o arquivo de contexto do cliente e retorna o texto."""
    arquivo = cliente.get("arquivo", "")
    context_path = SQUAD_ROOT / arquivo
    if not context_path.exists():
        return None, f"Arquivo não encontrado: {context_path}"
    with open(context_path, encoding="utf-8") as f:
        content = f.read().strip()
    # Remover linhas de comentário HTML
    content = re.sub(r"<!--.*?-->", "", content, flags=re.DOTALL).strip()
    if len(content) < 50:
        return None, "Arquivo de contexto está vazio ou muito curto"
    return content, None


# ---------------------------------------------------------------------------
# Actions
# ---------------------------------------------------------------------------

def action_detect_client(args):
    """Detecta automaticamente o cliente a partir do nome do arquivo xlsx."""
    clientes = load_client_index()
    if not clientes:
        print(json.dumps({
            "detectado": False,
            "motivo": "Nenhum cliente cadastrado em clients/_index.yaml",
            "clientes_disponiveis": []
        }, ensure_ascii=False))
        return

    cliente, confianca = detect_client_from_filename(args.file, clientes)

    if cliente and confianca == "alta":
        contexto, erro = load_client_context(cliente)
        print(json.dumps({
            "detectado": True,
            "confianca": "alta",
            "cliente": {
                "slug": cliente["slug"],
                "nome": cliente.get("nome", cliente["slug"]),
                "arquivo": cliente.get("arquivo"),
                "setor": cliente.get("setor"),
            },
            "contexto_disponivel": contexto is not None,
            "contexto_preview": contexto[:200] + "..." if contexto and len(contexto) > 200 else contexto,
            "contexto_completo": contexto,
            "erro_contexto": erro,
        }, ensure_ascii=False, indent=2))

    elif cliente and confianca == "baixa":
        print(json.dumps({
            "detectado": True,
            "confianca": "baixa",
            "aviso": "Múltiplos clientes corresponderam — selecionado o mais específico. Confirme se correto.",
            "cliente": {
                "slug": cliente["slug"],
                "nome": cliente.get("nome", cliente["slug"]),
            },
            "todos_os_clientes": [
                {"slug": c["slug"], "nome": c.get("nome", c["slug"])}
                for c in clientes if c.get("ativo", True)
            ],
        }, ensure_ascii=False, indent=2))

    else:
        print(json.dumps({
            "detectado": False,
            "motivo": "Nenhum cliente detectado pelo nome do arquivo",
            "arquivo_analisado": Path(args.file).name,
            "dica": "Inclua o slug do cliente no nome do arquivo (ex: mand-digital_posts.xlsx) ou use *set-empresa",
            "clientes_disponiveis": [
                {
                    "slug": c["slug"],
                    "nome": c.get("nome", c["slug"]),
                    "aliases": c.get("aliases", []),
                }
                for c in clientes if c.get("ativo", True)
            ],
        }, ensure_ascii=False, indent=2))


# ---------------------------------------------------------------------------
# Auto-onboarding helpers
# ---------------------------------------------------------------------------

def extract_company_name(context_text):
    """
    Tenta extrair o nome da empresa do texto de contexto.
    Estratégias em ordem de prioridade:
      1. Padrão explícito: "A {Nome} é uma empresa..." / "O {Nome} é uma..."
      2. Padrão de apresentação: "{Nome} é especializada..." / "{Nome} atua..."
      3. Primeira sequência de palavras capitalizadas nas primeiras 3 linhas
    Retorna (nome_extraido, confianca: 'alta'|'media'|'baixa')
    """
    # Limpar tags HTML/XML e comentários
    text = re.sub(r"<!--.*?-->", "", context_text, flags=re.DOTALL)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.strip()

    # Pegar as primeiras 500 chars para análise
    head = text[:500]

    # Padrão 1: "A/O {Nome Próprio} é uma empresa / é especializada / atua"
    patterns_alta = [
        r"\b[AO]s?\s+((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,5})\s+é\s+uma?\s+empresa",
        r"\b[AO]s?\s+((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,5})\s+é\s+especializada",
        r"\b[AO]s?\s+empresa\s+((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,5})\s+é",
    ]

    # Padrão 2: "{Nome} é uma empresa / é especializada / atua / foi fundada"
    patterns_media = [
        r"^((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,5})\s+é\s+uma?\s+empresa",
        r"^((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,5})\s+é\s+especializada",
        r"Fundada?\s+por\s+\w+[,.]?\s+[ao]\s+((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,4})\s+nasceu",
        r"((?:[A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜ][A-Za-zÁÉÍÓÚÂÊÎÔÛÃÕÇÀÜáéíóúâêîôûãõçàü]*\s*){1,5})\s+atua\s+",
    ]

    for pat in patterns_alta:
        m = re.search(pat, head, re.IGNORECASE | re.MULTILINE)
        if m:
            name = m.group(1).strip()
            if 2 <= len(name) <= 60:
                return name, "alta"

    for pat in patterns_media:
        m = re.search(pat, head, re.IGNORECASE | re.MULTILINE)
        if m:
            name = m.group(1).strip()
            if 2 <= len(name) <= 60:
                return name, "media"

    # Fallback: primeira linha não vazia com ≥2 palavras capitalizadas
    for line in text.splitlines()[:5]:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        words = line.split()
        cap_words = [w for w in words if w and w[0].isupper() and len(w) > 1]
        if 1 <= len(cap_words) <= 6:
            name = " ".join(cap_words[:5])
            if 2 <= len(name) <= 60:
                return name, "baixa"

    return None, "nenhuma"


def name_similarity(a, b):
    """
    Calcula similaridade entre dois nomes de empresa (0.0 a 1.0).
    Usa token overlap: proporção de tokens em comum.
    Sem dependências externas.
    """
    def tokenize(s):
        s = re.sub(r"[^a-z0-9\s]", "", slugify(s))
        tokens = set(s.split()) - {"de", "da", "do", "e", "a", "o", "s", "ltda", "me", "sa", "eireli"}
        return tokens

    ta = tokenize(a)
    tb = tokenize(b)
    if not ta or not tb:
        return 0.0
    inter = ta & tb
    union = ta | tb
    return len(inter) / len(union)  # Jaccard


def generate_aliases(slug, nome):
    """Gera aliases automáticos a partir do slug e nome."""
    aliases = set()
    aliases.add(slug)

    # Versões do slug
    aliases.add(slug.replace("-", ""))
    aliases.add(slug.replace("-", "_"))

    # Tokens individuais do slug (exceto palavras genéricas)
    stopwords = {"de", "da", "do", "e", "a", "o", "s", "ltda", "me", "sa", "eireli", "digital", "solutions", "tech"}
    tokens = [t for t in slug.split("-") if t and t not in stopwords and len(t) > 2]
    for t in tokens:
        aliases.add(t)

    # Iniciais se nome multi-palavra
    if nome:
        words = [w for w in nome.split() if w[0].isupper() and len(w) > 1]
        if len(words) >= 2:
            initials = "".join(w[0].lower() for w in words)
            if len(initials) >= 2:
                aliases.add(initials)

    return sorted(aliases)


def action_auto_onboard(args):
    """
    Primeiro acesso: recebe xlsx + texto de contexto da empresa.
    Extrai nome da empresa, verifica se já existe, e retorna:
      - use_existing: cliente já cadastrado (match exato ou alta similaridade)
      - confirm_match: cliente similar encontrado, precisa de confirmação
      - create_new: empresa nova, dados para criar
    """
    # Carregar contexto
    if not args.context:
        print(json.dumps({"error": "--context obrigatório (path para arquivo de texto ou use --context-inline)"}))
        sys.exit(1)

    ctx_path = Path(args.context)
    if ctx_path.exists():
        with open(ctx_path, encoding="utf-8") as f:
            context_text = f.read()
    else:
        # Pode ter sido passado o texto diretamente via --context
        context_text = args.context

    # Remover tags <contexto-empresa> se presentes
    context_text = re.sub(r"</?contexto-empresa[^>]*>", "", context_text).strip()

    if len(context_text) < 50:
        print(json.dumps({"error": "Texto de contexto muito curto (< 50 chars)"}))
        sys.exit(1)

    # Extrair nome da empresa
    nome_extraido, confianca_nome = extract_company_name(context_text)

    # Carregar clientes existentes
    clientes = load_client_index()

    # 1. Tentar detecção pelo nome do arquivo xlsx primeiro
    cliente_por_filename, conf_filename = detect_client_from_filename(
        args.file or "", clientes
    ) if args.file else (None, "nenhuma")

    # 2. Tentar match pelo nome extraído do contexto
    melhor_match = None
    melhor_similaridade = 0.0

    if nome_extraido:
        for c in clientes:
            sim_nome = name_similarity(nome_extraido, c.get("nome", ""))
            sim_slug = name_similarity(nome_extraido, c.get("slug", "").replace("-", " "))
            sim = max(sim_nome, sim_slug)
            if sim > melhor_similaridade:
                melhor_similaridade = sim
                melhor_match = c

    # Decidir ação
    slug_sugerido = slugify(nome_extraido) if nome_extraido else None
    aliases_sugeridos = generate_aliases(slug_sugerido, nome_extraido) if slug_sugerido else []

    # Match exato pelo filename (alta confiança)
    if cliente_por_filename and conf_filename == "alta":
        ctx_existente, _ = load_client_context(cliente_por_filename)
        print(json.dumps({
            "acao": "use_existing",
            "origem_deteccao": "filename",
            "cliente": {
                "slug": cliente_por_filename["slug"],
                "nome": cliente_por_filename.get("nome"),
                "arquivo": cliente_por_filename.get("arquivo"),
            },
            "contexto_atual_disponivel": ctx_existente is not None,
            "nome_extraido_do_contexto": nome_extraido,
            "similaridade": 1.0,
            "mensagem": f"Cliente '{cliente_por_filename.get('nome')}' já está cadastrado e foi detectado pelo nome do arquivo.",
            "pergunta": "Quer atualizar o contexto com o texto fornecido agora? [S/N]",
        }, ensure_ascii=False, indent=2))
        return

    # Match pelo nome com alta similaridade (≥ 0.7)
    if melhor_match and melhor_similaridade >= 0.7:
        ctx_existente, _ = load_client_context(melhor_match)
        print(json.dumps({
            "acao": "use_existing",
            "origem_deteccao": "nome_similaridade",
            "cliente": {
                "slug": melhor_match["slug"],
                "nome": melhor_match.get("nome"),
                "arquivo": melhor_match.get("arquivo"),
            },
            "contexto_atual_disponivel": ctx_existente is not None,
            "nome_extraido_do_contexto": nome_extraido,
            "similaridade": round(melhor_similaridade, 2),
            "mensagem": f"Parece que este contexto é de '{melhor_match.get('nome')}' (já cadastrado).",
            "pergunta": "Confirma que é o mesmo cliente? [S/N] — Se não, o sistema criará um novo cadastro.",
        }, ensure_ascii=False, indent=2))
        return

    # Match parcial (0.3–0.69) — pede confirmação
    if melhor_match and melhor_similaridade >= 0.3:
        print(json.dumps({
            "acao": "confirm_match",
            "nome_extraido_do_contexto": nome_extraido,
            "confianca_extracao": confianca_nome,
            "cliente_similar": {
                "slug": melhor_match["slug"],
                "nome": melhor_match.get("nome"),
                "similaridade": round(melhor_similaridade, 2),
            },
            "slug_sugerido_para_novo": slug_sugerido,
            "aliases_sugeridos": aliases_sugeridos,
            "mensagem": f"Encontrei '{melhor_match.get('nome')}' já cadastrado, mas a similaridade com '{nome_extraido}' é baixa ({round(melhor_similaridade*100)}%).",
            "pergunta": f"Este contexto é da mesma empresa '{melhor_match.get('nome')}'? [S] Sim (usa existente) / [N] Não (cria '{nome_extraido}')",
        }, ensure_ascii=False, indent=2))
        return

    # Empresa nova — retornar dados para criação
    print(json.dumps({
        "acao": "create_new",
        "nome_extraido": nome_extraido,
        "confianca_extracao": confianca_nome,
        "slug_sugerido": slug_sugerido,
        "aliases_sugeridos": aliases_sugeridos,
        "contexto_texto": context_text,
        "mensagem": f"Empresa '{nome_extraido}' não encontrada nos clientes cadastrados. Pronta para criar.",
        "instrucao": "Use --action=save-client para criar o cadastro com estes dados.",
    }, ensure_ascii=False, indent=2))


def action_save_client(args):
    """
    Cria/atualiza o arquivo clients/{slug}.md e a entrada em _index.yaml.
    Usado após auto-onboard retornar 'create_new' ou confirmação de update.
    """
    if not args.slug or not args.nome:
        print(json.dumps({"error": "--slug e --nome obrigatórios"}))
        sys.exit(1)

    # Carregar contexto
    context_text = ""
    if args.context:
        ctx_path = Path(args.context)
        if ctx_path.exists():
            with open(ctx_path, encoding="utf-8") as f:
                context_text = f.read()
        else:
            context_text = args.context
        context_text = re.sub(r"</?contexto-empresa[^>]*>", "", context_text).strip()

    slug = slugify(args.slug)
    nome = args.nome.strip()
    setor = args.setor.strip() if args.setor else ""
    aliases_raw = args.aliases.split(",") if args.aliases else []
    aliases = list(set(generate_aliases(slug, nome) + [a.strip() for a in aliases_raw if a.strip()]))
    aliases = sorted(set(aliases))

    # 1. Criar/atualizar clients/{slug}.md
    client_file = SQUAD_ROOT / "clients" / f"{slug}.md"
    client_file.parent.mkdir(parents=True, exist_ok=True)

    md_content = f"# Contexto da Empresa — {nome}\n\n{context_text}\n"
    with open(client_file, "w", encoding="utf-8") as f:
        f.write(md_content)

    # 2. Atualizar _index.yaml
    index_path = SQUAD_ROOT / "clients" / "_index.yaml"
    clientes = load_client_index()

    # Verificar se slug já existe
    exists = any(c["slug"] == slug for c in clientes)
    action_taken = "atualizado" if exists else "criado"

    if index_path.exists():
        with open(index_path, "r", encoding="utf-8") as f:
            index_content = f.read()

        if exists:
            # Remover entrada existente (bloco simples entre "- slug: {slug}" e o próximo "- slug:")
            pattern = rf"(  - slug: {re.escape(slug)}\n(?:    [^\n]*\n)*)"
            index_content = re.sub(pattern, "", index_content)

        # Montar nova entrada
        aliases_yaml = "\n".join(f"      - {a}" for a in aliases)
        setor_yaml = f'\n    setor: "{setor}"' if setor else ""
        new_entry = (
            f"\n  - slug: {slug}\n"
            f'    nome: "{nome}"\n'
            f'    arquivo: "clients/{slug}.md"\n'
            f"    aliases:\n{aliases_yaml}{setor_yaml}\n"
            f"    ativo: true\n"
        )

        # Inserir antes dos comentários de template (se houver) ou no final da lista
        if "# Template para novos clientes" in index_content:
            index_content = index_content.replace(
                "  # Template para novos clientes", new_entry + "\n  # Template para novos clientes"
            )
        elif "clientes:" in index_content:
            # Append antes do fim do arquivo
            index_content = index_content.rstrip() + "\n" + new_entry
        else:
            index_content += "\nclientes:" + new_entry

        with open(index_path, "w", encoding="utf-8") as f:
            f.write(index_content)
    else:
        # Criar _index.yaml do zero
        aliases_yaml = "\n".join(f"      - {a}" for a in aliases)
        setor_yaml = f'\n    setor: "{setor}"' if setor else ""
        with open(index_path, "w", encoding="utf-8") as f:
            f.write(
                f"clientes:\n"
                f"  - slug: {slug}\n"
                f'    nome: "{nome}"\n'
                f'    arquivo: "clients/{slug}.md"\n'
                f"    aliases:\n{aliases_yaml}{setor_yaml}\n"
                f"    ativo: true\n"
            )

    print(json.dumps({
        "sucesso": True,
        "acao": action_taken,
        "slug": slug,
        "nome": nome,
        "arquivo": str(client_file),
        "aliases": aliases,
        "contexto_tamanho": len(context_text),
        "mensagem": f"Cliente '{nome}' {action_taken} com sucesso.",
        "dica_nome_arquivo": f"Nomeie seus xlsx com '{slug}' para detecção automática. Ex: {slug}_posts.xlsx",
    }, ensure_ascii=False, indent=2))


def action_list_clients(args):
    """Lista todos os clientes cadastrados e status do arquivo de contexto."""
    clientes = load_client_index()
    result = []
    for c in clientes:
        context_path = SQUAD_ROOT / c.get("arquivo", "")
        exists = context_path.exists()
        size = context_path.stat().st_size if exists else 0
        result.append({
            "slug": c["slug"],
            "nome": c.get("nome", c["slug"]),
            "arquivo": c.get("arquivo"),
            "setor": c.get("setor"),
            "ativo": c.get("ativo", True),
            "contexto_existe": exists,
            "contexto_tamanho_bytes": size,
            "aliases": c.get("aliases", []),
        })
    print(json.dumps({
        "total": len(result),
        "clientes": result,
        "index_path": str(SQUAD_ROOT / "clients" / "_index.yaml"),
    }, ensure_ascii=False, indent=2))


def action_read(args):
    """Lê o XLSX, detecta cliente e exibe as linhas processáveis."""
    openpyxl = get_openpyxl()
    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"Arquivo não encontrado: {path}"}))
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(json.dumps({"error": f"Não foi possível abrir o arquivo: {e}"}))
        sys.exit(1)

    ws = wb.active
    col_map, col_names = read_headers(ws)

    if "keyword" not in col_map:
        print(json.dumps({
            "error": "Coluna 'Palavra-chave Foco' não encontrada na linha 1.",
            "headers_found": list(col_names.values())
        }))
        sys.exit(1)
    if "titulo" not in col_map:
        print(json.dumps({
            "error": "Coluna 'Título' não encontrada na linha 1.",
            "headers_found": list(col_names.values())
        }))
        sys.exit(1)

    # Detectar cliente automaticamente
    clientes = load_client_index()
    cliente_detectado, confianca = detect_client_from_filename(args.file, clientes)
    cliente_info = None
    contexto_cliente = None

    if cliente_detectado:
        ctx, err = load_client_context(cliente_detectado)
        cliente_info = {
            "slug": cliente_detectado["slug"],
            "nome": cliente_detectado.get("nome"),
            "confianca": confianca,
        }
        if ctx:
            contexto_cliente = ctx

    # Ler linhas
    processable = []
    already_processed = []
    incomplete = []

    for row_idx in range(2, ws.max_row + 1):
        def get_val(field, ridx=row_idx):
            col = col_map.get(field)
            if col is None:
                return None
            val = ws.cell(row=ridx, column=col).value
            return val if val not in (None, "") else None

        keyword = get_val("keyword")
        titulo = get_val("titulo")
        idioma = get_val("idioma")
        categoria = get_val("categoria")
        prompt_atual = get_val("prompt_adicional")

        if keyword is None and titulo is None:
            continue

        row_data = {
            "numero": row_idx,
            "keyword": str(keyword) if keyword else None,
            "titulo": str(titulo) if titulo else None,
            "idioma": str(idioma) if idioma else "Português (Brasil)",
            "categoria": str(categoria) if categoria else None,
            "prompt_preenchido": prompt_atual is not None,
        }

        if keyword and titulo:
            if prompt_atual is None:
                processable.append(row_data)
            else:
                already_processed.append(row_data)
        else:
            incomplete.append(row_data)

    print(json.dumps({
        "arquivo": str(path),
        "cliente_detectado": cliente_info,
        "contexto_cliente": contexto_cliente,
        "total_linhas": len(processable) + len(already_processed) + len(incomplete),
        "processaveis": processable,
        "ja_processadas": already_processed,
        "incompletas": incomplete,
    }, ensure_ascii=False, indent=2))


def action_write(args):
    """Escreve briefings nas colunas Prompt Adicional e Termos LSI."""
    openpyxl = get_openpyxl()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(json.dumps({"error": f"Arquivo de entrada não encontrado: {input_path}"}))
        sys.exit(1)

    if args.data:
        data_path = Path(args.data)
        if data_path.exists():
            with open(data_path, encoding="utf-8") as f:
                results = json.load(f)
        elif args.data.lstrip().startswith("[") or args.data.lstrip().startswith("{"):
            results = json.loads(args.data)
        else:
            print(json.dumps({"error": f"--data não é um path válido nem JSON: {args.data}"}))
            sys.exit(1)
    else:
        print(json.dumps({"error": "Dados de resultados não fornecidos (--data)"}))
        sys.exit(1)

    if isinstance(results, dict) and "linhas" in results:
        results = results["linhas"]

    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(json.dumps({"error": f"Não foi possível abrir o arquivo: {e}"}))
        sys.exit(1)

    ws = wb.active
    col_map, _ = read_headers(ws)

    if "prompt_adicional" not in col_map:
        print(json.dumps({"error": "Coluna 'Prompt Adicional' não encontrada na planilha"}))
        sys.exit(1)
    if "termos_lsi" not in col_map:
        print(json.dumps({"error": "Coluna 'Termos LSI' não encontrada na planilha"}))
        sys.exit(1)

    col_prompt = col_map["prompt_adicional"]
    col_lsi = col_map["termos_lsi"]
    written = []
    errors = []

    for item in results:
        row_num = item.get("numero")
        if row_num is None:
            errors.append({"item": str(item)[:80], "erro": "Campo 'numero' ausente"})
            continue
        try:
            prompt = item.get("prompt_adicional", "")
            lsi = item.get("termos_lsi", "")
            if prompt:
                ws.cell(row=row_num, column=col_prompt).value = str(prompt)
            if lsi:
                ws.cell(row=row_num, column=col_lsi).value = str(lsi)
            written.append({"linha": row_num, "prompt_chars": len(str(prompt)), "lsi_chars": len(str(lsi))})
        except Exception as e:
            errors.append({"linha": row_num, "erro": str(e)})

    # Nunca sobrescrever o original
    if output_path.resolve() == input_path.resolve():
        output_path = input_path.parent / (input_path.stem + "_contextualizado" + input_path.suffix)

    try:
        wb.save(output_path)
    except Exception as e:
        print(json.dumps({"error": f"Não foi possível salvar: {e}"}))
        sys.exit(1)

    print(json.dumps({
        "sucesso": True,
        "arquivo_saida": str(output_path),
        "linhas_escritas": len(written),
        "detalhes": written,
        "erros": errors,
    }, ensure_ascii=False, indent=2))


def action_list_columns(args):
    openpyxl = get_openpyxl()
    path = Path(args.file)
    if not path.exists():
        print(json.dumps({"error": f"Arquivo não encontrado: {path}"}))
        sys.exit(1)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    columns = []
    for cell in ws[1]:
        if cell.value is not None:
            columns.append({
                "col_letra": cell.column_letter,
                "col_numero": cell.column,
                "header": str(cell.value),
                "mapeado_para": HEADER_MAP.get(normalize(str(cell.value)), "NÃO MAPEADO"),
            })
    print(json.dumps({"arquivo": str(path), "total_colunas": len(columns), "colunas": columns},
                     ensure_ascii=False, indent=2))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="SEO Contextualizador — XLSX Manager")
    parser.add_argument("--action",
                        choices=["read", "write", "detect-client", "auto-onboard",
                                 "save-client", "list-clients", "list-columns"],
                        required=True)
    parser.add_argument("--file",    help="Arquivo XLSX (read, detect-client, auto-onboard, list-columns)")
    parser.add_argument("--input",   help="Arquivo XLSX de entrada (write)")
    parser.add_argument("--output",  help="Arquivo XLSX de saída (write)")
    parser.add_argument("--data",    help="JSON com resultados a escrever — path ou string JSON")
    parser.add_argument("--context", help="Texto de contexto da empresa — path para .txt ou string")
    parser.add_argument("--slug",    help="Slug do cliente (save-client)")
    parser.add_argument("--nome",    help="Nome do cliente (save-client)")
    parser.add_argument("--setor",   help="Setor/nicho do cliente (save-client)")
    parser.add_argument("--aliases", help="Aliases extras separados por vírgula (save-client)")
    args = parser.parse_args()

    if args.action == "read":
        if not args.file:
            print(json.dumps({"error": "--file obrigatório"})); sys.exit(1)
        action_read(args)

    elif args.action == "write":
        if not args.input or not args.output:
            print(json.dumps({"error": "--input e --output obrigatórios"})); sys.exit(1)
        action_write(args)

    elif args.action == "detect-client":
        if not args.file:
            print(json.dumps({"error": "--file obrigatório"})); sys.exit(1)
        action_detect_client(args)

    elif args.action == "auto-onboard":
        if not args.context:
            print(json.dumps({"error": "--context obrigatório"})); sys.exit(1)
        action_auto_onboard(args)

    elif args.action == "save-client":
        if not args.slug or not args.nome:
            print(json.dumps({"error": "--slug e --nome obrigatórios"})); sys.exit(1)
        action_save_client(args)

    elif args.action == "list-clients":
        action_list_clients(args)

    elif args.action == "list-columns":
        if not args.file:
            print(json.dumps({"error": "--file obrigatório"})); sys.exit(1)
        action_list_columns(args)


if __name__ == "__main__":
    main()
